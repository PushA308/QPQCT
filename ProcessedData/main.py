import os
import sys
import traceback
import win32com.client #pip install pywin32
import csv
import nltk
import shutil
import xlsxwriter
import openpyxl
#import matplotlib.pyplot as plt; plt.rcdefaults()
import numpy as np
import matplotlib.pyplot as plt
#nltk.download('punkt')
#nltk.download('averaged_perceptron_tagger')
#include other ntlk packages, if asked for.

##########################################
#initialized variables
##########################################
question_no_column = 7
start_index = 16
mainDir = "C:\\xampp\\htdocs\\www\\WebPage\\BTechProject\\"


def compare_co_and_questiontype():
    f1=open("new_co_file.csv","r")
    f2=open(mainDir+"new_bloom_type.csv","r")
    fd = open(os.path.join(os.getcwd(),sys.argv[1]+"analysed_co_and_questiontype.csv"),'w')
    r1 = list(f1)
    length = len(r1)
    r2 = list(f2)
    sum=0
    for i in range(length):
        if r1[i]==r2[i]:
            k=abs(int(r1[i])-int(r2[i]))
            sum=sum+k
            fd.write((chr(ord('A') + k))+'\n')
        else:
            k=abs(int(r1[i])-int(r2[i]))
            sum=sum+k
            fd.write((chr(ord('A') + k))+'\n')
    avg_quality_question_paper_using_co_and_questiontype=chr(ord('A')+int(sum/length))
    return avg_quality_question_paper_using_co_and_questiontype



def compare_cowise():
    f1=open("new_co_file.csv","r")
    f2=open(mainDir+"File\\Value_level.csv","r")
    fd = open(os.path.join(os.getcwd(),sys.argv[1]+"analysed_CO.csv"),'w')
    r1 = list(f1)
    length = len(r1) - 1
    r2 = list(f2)
    sum=0
    for i in range(length):
        if r1[i]==r2[i]:
            k=abs(int(r1[i])-int(r2[i]))
            sum=sum+k
            fd.write((chr(ord('A') + k))+'\n')
        else:
            k=abs(int(r1[i])-int(r2[i]))
            sum=sum+k
            fd.write((chr(ord('A') + k))+'\n')
    avg_quality_question_paper_using_cowise=(chr(ord('A')+int(sum/length)))
    return avg_quality_question_paper_using_cowise




def generate_COFile(CO):
	f = open(sys.argv[1]+"processed_data.csv","r")
	reader=csv.reader(f)
	fd = open(os.path.join(os.getcwd(),"new_co_file.csv"),'w')
	included_cols=[1]
	for row in reader:
		content=list(row[i] for i in included_cols)
		if content[0]=="CO1":
			fd.write(str(1)+'\n')
		elif content[0]=="CO2":
			fd.write(str(2)+'\n')
		elif content[0]=="CO3":
			fd.write(str(3)+'\n')



   
	
	
	
def compare_QuestionType():
    f1=open("new_bloom_type.csv","r")
    f2=open(mainDir+"File\\Value_level.csv","r")
    fd = open(os.path.join(os.getcwd(),sys.argv[1]+"analysed_questiontype.csv"),'w')
    r1 = list(f1)
    r2 = list(f2)
    length = len(r2) 
    sum=0
    for i in range(length):
        if r1[i]==r2[i]:
            k=abs(int(r1[i])-int(r2[i]))
            sum=sum+k
            fd.write((chr(ord('A') + k))+'\n')
        else:
            k=abs(int(r1[i])-int(r2[i]))
            sum=sum+k
            fd.write((chr(ord('A') + k))+'\n')
    avg_quality=chr(ord('A')+int(sum/length))
    return avg_quality

def Bloom_Type():
    f = open(sys.argv[1]+"processed_data.csv","r")
    reader = csv.reader(f)
    #out_file = open("File\solution1.csv", "w")
    #writer = csv.writer(out_file)
    add=0
    included_cols = [2]
    included_cols1=[3]
    fd = open(os.path.join(os.getcwd(),mainDir+"new_bloom_type.csv"),'w')
    for row in reader:
       content = list(row[i] for i in included_cols) #selecting question
       content1 = list(row[i] for i in included_cols1) #selecting question type
       if content1[0]=="Drg":
           fd.write("1"+'\n')
       elif content1[0]=="An-S":
           fd.write("2"+'\n')
       elif content1[0]=="Desc":
           fd.write("3"+'\n')
       elif content1[0]=="Des":
           fd.write("4"+'\n')
       elif content1[0]=="Prog":
           fd.write("5"+'\n')
       elif content1[0]=="Obj":
           fd.write("6"+'\n')
       else:
           fd.write("0"+'\n')

def final(k):
    #fs = open("File/level.csv","r")
    fs = open(mainDir+"File\\level.csv","r")
    reader = csv.reader(fs)
    i=0
    res=0
    b=0
    d=0
    fd = open(os.path.join(os.getcwd(),mainDir+"File\\Value_level.csv"),'w')
    for row in reader:
    	fd.write(max(row)+'\n')
    	res=res+int(max(row))
    	i=i+1
    fs.close()
    if k==1:
        p=res
        print(i)
        print(p)
        l=6
        k=(p/(i*l))*100
        res_value='F'
        if k>=90:
        	res_value='A'
        elif k>=80:
        	res_value='B'
        elif k>=70:
        	res_value='C'
        elif k>=60:
        	res_value='D'
        elif k>=50:
        	res_value='E'
        fp = open(os.path.join(os.getcwd(),"result_verb.csv"),'w')
        fp.write(res_value)
        print("Average quality per question is {}".format(k))
        fa = open(os.path.join(os.getcwd(),sys.argv[1]+"Value.csv"),'w')
        fa.write(str(res_value))
        fs.close()

def calculate_level(line):
    #fs = open(mainDir+"File\\bloom verbs.csv","r")
    #reader = csv.reader(fs)
    #included_cols = [1]
    a=list()
    b=list()
    #row = next(reader)
    #print(line)
    flg=0
    for word in line:
        i=1
        fs = open(mainDir+"File\\bloom verbs.csv","r")
        reader = csv.reader(fs)
        for row in reader:
            if word.lower() in row:
                a.append(i)
                flg=1
            i=i+1
        fs.close()
        if flg==0:
            a.append(0)
        #print(line,a,max(a))
    with open(mainDir+"File\\level.csv","a",newline='') as csvfile:
            spamwriter = csv.writer(csvfile)
            spamwriter.writerow(a)
    fs.close()

def extract_verb(sentence):
    helping_verbs = ['am','are','is','was','were','be','being','been','have','has','had','shall','will','do','does','did','may','might','must','can','could','would','should']
    #sentence = "what is meant by data structure. it has been long time. i didn't do that"
    tokens = nltk.word_tokenize(sentence)
    tagged = nltk.pos_tag(tokens)
    #print("tagged tokens:-")
    #print(tagged)
    length = len(tagged) - 1
    #print(length)
    a = list()
    i=0
    flg=0
    for item in tagged:
        if item[1][0] == 'V':
            if((item[0] in helping_verbs)!=1):
                a.append(item[0])
                #print(item[0])
                flg=1
    if flg==0:
        a.append("N.A")
    #print(a)
    with open(sys.argv[1]+"question_verb.csv","a",newline='') as csvfile:
            spamwriter = csv.writer(csvfile)
            spamwriter.writerow(a)
    return a;
    
def process_question_paper(ques_paper_path) :
    fd = open(os.path.join(os.getcwd(),sys.argv[1]+"processed_data.csv"),'w')
    for root, dirs,files in os.walk(ques_paper_path) :
        for file in files:
            if file.endswith('.xls') :
                file_path = os.path.join(root,file)
                try:
                    excel = win32com.client.Dispatch('Excel.Application')
                    workbook = excel.Workbooks.open(file_path)
                    sheet = workbook.WorkSheets('QuestionPaper')
                    for start_row in range(start_index, 50):
                        try:
                            row, col = start_row, question_no_column
                            question = sheet.Cells(row, col).value

                            if question is not None:
                                row, col = start_row, question_no_column + 1
                                marks = str(sheet.Cells(row, col).value)
                                row, col = start_row, question_no_column + 2
                                co_type = str(sheet.Cells(row, col).value)
                                row, col = start_row, question_no_column + 4
                                module_no = str(sheet.Cells(row, col).value)
                                row, col = start_row, question_no_column - 5
                                question_type =  sheet.Cells(row, col).value
                                row, col = start_row, question_no_column - 2
                                question_no =  sheet.Cells(row, col).value
                                row, col = start_row, question_no_column - 1
                                sub_question_no = sheet.Cells(row, col).value
                                row, col = start_row, question_no_column - 2
                                question_no =  sheet.Cells(row, col).value
                                row, col = start_row, question_no_column
                                question =  sheet.Cells(row, col).value
                                a=extract_verb(question)
                                print(a)
                                calculate_level(a)
                                tokens = nltk.word_tokenize(question.lower())
                                text = nltk.Text(tokens)
                                verbs_list = [wrd for (wrd, tags) in nltk.pos_tag(text) if tags in ('VBG')]
                                fd.write(marks + ','+co_type + ',' + module_no + ',' +question_type + ','+ question_no + ',' + sub_question_no + ',' +' | '.join(verbs_list)+ '\n')
                            else:
                                pass


                        except Exception as e:
                            print ("hhj")
                            pass
                    workbook.Close(True)
                except Exception as e:
                    print ("ERROR")
                    pass
    fd.close()


filename = mainDir+"File\\level.csv"
filename2 = sys.argv[1]+"question_verb.csv"
if os.path.exists(filename): os.remove(filename)
if os.path.exists(filename2): os.remove(filename2)
if __name__ == "__main__" :
    arg_cnt = len(sys.argv)
    if arg_cnt > 1:
        ques_paper_path = sys.argv[1]
        process_question_paper(ques_paper_path)
    else:
        print ("Please provide question paper directory path !")
#Verbs_Separator()
Bloom_Type()
z=sys.argv[1]+"processed_data.csv"	
wb = xlsxwriter.Workbook(z.replace(".csv",".xlsx"))
ws = wb.add_worksheet("WS1")    # your worksheet title here
with open(z,'r') as csvfile:
    table = csv.reader(csvfile)
    i = 0
    # write each row from the csv file as text into the excel file
    # this may be adjusted to use 'excel types' explicitly (see xlsxwriter doc)
    for row in table:
        ws.write_row(i, 0, row)
        i += 1
wb.close()
	
excel_document=openpyxl.load_workbook(sys.argv[1]+'processed_data.xlsx')
sheet = excel_document['WS1']
ws = excel_document.active
first_column = ws['D']
mark_column=ws['A']
module_column=ws['C']
a=list()
# Print the contents
module=list();
for i in range(len(module_column)):
	if module_column[i].value not in module:
		module.append(module_column[i].value);
print(module)
chapterwise_mark=list()
for i in range(len(module)):
    chapter=module[i]
    k=0
    for j in range(len(module_column)):
        if chapter == module_column[j].value:
            z=float(mark_column[j].value)
            t=int(z)
            k=k+t
    chapterwise_mark.append(k)	
print(chapterwise_mark)
#---------------------------------------------------------------------------

#barplot of marks and module

y_pos = np.arange(len(module))
 
plt.bar(y_pos, chapterwise_mark, width=0.5, align='center', alpha=0.5)
plt.xticks(y_pos, module)
plt.ylabel('Marks')
plt.title('Modulewise Marks')
 
#plt.show()
plt.savefig(sys.argv[1]+"1.png")
plt.close()
#-----------------------------------------------------------------------------
CO_column=ws['B']
CO=list()
for i in range(len(CO_column)):
	if CO_column[i].value not in CO:
		CO.append(CO_column[i].value);
print(CO)

cowise_mark=list()
for i in range(len(CO)):
    co=CO[i]
    k=0
    for j in range(len(CO_column)):
        if co == CO_column[j].value:
            z=float(mark_column[j].value)
            t=int(z)
            k=k+t
    cowise_mark.append(k)	
print(cowise_mark)

#--------------------------------------------------------------------

#barplot of CO and mark

y_pos = np.arange(len(CO))
 
plt.bar(y_pos, cowise_mark,width=0.5, align='center', alpha=0.5)
plt.xticks(y_pos, CO)
plt.ylabel('Marks')
plt.title('COwise Marks')
 
#plt.show()
plt.savefig(sys.argv[1]+"2.png")
plt.close()
#-------------------------------------------------------------------------
QueType_column=ws['D']
Type=list()
for i in range(len(QueType_column)):
	if QueType_column[i].value not in Type:
		Type.append(QueType_column[i].value);
print(Type)

Questiontypewise_mark=list()
for i in range(len(Type)):
    type=Type[i]
    k=0
    for j in range(len(QueType_column)):
        if type == QueType_column[j].value:
           z=float(mark_column[j].value)
           t=int(z)
           k=k+t
    Questiontypewise_mark.append(k)	
print(Questiontypewise_mark)

#-------------------------------------------------------------------------
#barplot of Question Type and mark

y_pos = np.arange(len(Type))
 
plt.bar(y_pos, Questiontypewise_mark, width=0.5, align='center', alpha=0.5)
plt.xticks(y_pos, Type)
plt.ylabel('Marks')
plt.title('Question Typewise Marks')
 
#plt.show()
plt.savefig(sys.argv[1]+"3.png")
plt.close()
#--------------------------------------------------------------------------


generate_COFile(CO)
#k=int(input("Select the option for Analysis of Question Paper:\n1.Verbs\n2.Question Type\n3.Course Outcome"))
k = int(sys.argv[2])
if k==1:
    final(k)
elif k==2:
	final(k)
	avg_quality=compare_QuestionType()
	print("Avg Quality:"+ avg_quality)
	fd = open(os.path.join(os.getcwd(),sys.argv[1]+"result_questiontype.csv"),'w')
	fd.write(avg_quality)
	fd.close()

elif k==3:
	final(k)
	avg_quality=compare_cowise()
	print("Avg Quality:"+ avg_quality)
	fd = open(os.path.join(os.getcwd(),sys.argv[1]+"result_cowise.csv"),'w')
	fd.write(avg_quality)
	fd.close()
elif k==4:
	final(k)
	avg_quality=compare_co_and_questiontype()
	print("Avg Quality:"+ avg_quality)
	fd = open(os.path.join(os.getcwd(),sys.argv[1]+"result_co_and_questiontype.csv"),'w')
	fd.write(avg_quality)
	fd.close()