import openpyxl
import csvimport nltkimport shutil
import os
import matplotlib.pyplot as plt; plt.rcdefaults()
import numpy as np
import matplotlib.pyplot as plt
def extract_verb(sentence):    helping_verbs = ['am','are','is','was','were','be','being','been','have','has','had','shall','will','do','does','did','may','might','must','can','could','would','should']    #sentence = "what is meant by data structure. it has been long time. i didn't do that"    tokens = nltk.word_tokenize(sentence)    tagged = nltk.pos_tag(tokens)    #print("tagged tokens:-")    #print(tagged)    length = len(tagged) - 1    #print(length)    a = list()    i=0    flg=0    for item in tagged:        if item[1][0] == 'V':            if((item[0] in helping_verbs)!=1):                a.append(item[0])                #print(item[0])                flg=1    if flg==0:        a.append("N.A")    #print(a)
    #os.remove("NewVerbs.csv");    with open("NewVerbs.csv","a",newline='') as csvfile:            spamwriter = csv.writer(csvfile)            spamwriter.writerow(a)    return a;
	
def calculate_level(line):
    #fs = open("File/bloom verbs.csv","r")
    #reader = csv.reader(fs)
    #included_cols = [1]
    a=list()
    b=list()
    #row = next(reader)
    #print(line)
    flg=0
    for word in line:
        fs = open("File/bloom verbs.csv","r")
        reader = csv.reader(fs)
        i=1
        for row in reader:
            if word.lower() in row:
                a.append(i)
                flg=1
            i=i+1
        if flg==0:
            a.append(0)
        #print(line,a,max(a))
    with open("New_level.csv","a",newline='') as csvfile:
            spamwriter = csv.writer(csvfile)
            spamwriter.writerow(a)
    fs.close()
	excel_document=openpyxl.load_workbook('UOS1.xlsx')
sheet = excel_document.get_sheet_by_name('Sheet1')
ws = excel_document.active
first_column = ws['D']
mark_column=ws['E']
module_column=ws['G']
a=list()
# Print the contents
for x in range(len(first_column)): 
    a.append(first_column[x].value) for i in range(len(a)):	a[i]=a[i].replace('=',' ')for sentence in a:	a=extract_verb(sentence)
	calculate_level(a)
module=list();
for i in range(len(module_column)):
	if module_column[i].value not in module:
		module.append(module_column[i].value);
print(module)
chapterwise_mark=list()
for i in range(len(module)):
    chapter=module[i]
    k=0
    for j in range(len(module_column)):
        if chapter == module_column[j].value:
            k=k+mark_column[j].value
    chapterwise_mark.append(k)	
print(chapterwise_mark)
#---------------------------------------------------------------------------

#barplot of marks and module

y_pos = np.arange(len(module))
 
plt.bar(y_pos, chapterwise_mark, align='center', alpha=0.5)
plt.xticks(y_pos, module)
plt.ylabel('Marks')
plt.title('Modulewise Marks')
 
plt.show()
#-----------------------------------------------------------------------------
CO_column=ws['F']
CO=list()
for i in range(len(CO_column)):
	if CO_column[i].value not in CO:
		CO.append(CO_column[i].value);
print(CO)

cowise_mark=list()
for i in range(len(CO)):
    co=CO[i]
    k=0
    for j in range(len(CO_column)):
        if co == CO_column[j].value:
            k=k+mark_column[j].value
    cowise_mark.append(k)	
print(cowise_mark)

#--------------------------------------------------------------------

#barplot of CO and mark

y_pos = np.arange(len(CO))
 
plt.bar(y_pos, cowise_mark, align='center', alpha=0.5)
plt.xticks(y_pos, CO)
plt.ylabel('Marks')
plt.title('COwise Marks')
 
plt.show()

#-------------------------------------------------------------------------
QueType_column=ws['A']
Type=list()
for i in range(len(QueType_column)):
	if QueType_column[i].value not in Type:
		Type.append(QueType_column[i].value);
print(Type)

Questiontypewise_mark=list()
for i in range(len(Type)):
    type=Type[i]
    k=0
    for j in range(len(QueType_column)):
        if type == QueType_column[j].value:
            k=k+mark_column[j].value
    Questiontypewise_mark.append(k)	
print(Questiontypewise_mark)

#-------------------------------------------------------------------------
#barplot of Question Type and mark

y_pos = np.arange(len(Type))
 
plt.bar(y_pos, Questiontypewise_mark, align='center', alpha=0.5)
plt.xticks(y_pos, Type)
plt.ylabel('Marks')
plt.title('Question Typewise Marks')
 
plt.show()

#--------------------------------------------------------------------------
